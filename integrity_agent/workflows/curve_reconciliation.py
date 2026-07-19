from __future__ import annotations

import csv
import hashlib
import io
import json
import math
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from integrity_agent.core.curves import (
    CurveColumnMapping,
    CurvePoint,
    CurveReconciliationSpec,
    CurveSegmentSimilarityOptions,
    CurveTableSpec,
)
from integrity_agent.core.evidence.schema import (
    EvidenceItem,
    Finding,
    ManualVerification,
    RiskLevel,
)
from integrity_agent.core.rules.registry import load_rule_registry
from integrity_agent.core.rules.schema import DetectorRule, RuleExecutionResult
from integrity_agent.workflows.validate_ledger import validate_ledger_file


RULE_ID = "curve_point_coverage"
SEGMENT_SIMILARITY_RULE_ID = "curve_segment_shape_similarity"
DEFAULT_OUTPUT_DIR = Path("outputs") / "curve_reconciliation"
FINDINGS_NAME = "curve_point_coverage.jsonl"
SUMMARY_NAME = "curve_source_reconciliation_summary.md"
SAFE_MISSING_INTERVAL_LANGUAGE = (
    "Candidate curve-coverage question: a contiguous internal x-axis interval "
    "from the supplied source table is not represented in the supplied plot-data "
    "table; verify plotting context and disclosed processing."
)
SAFE_CONTEXT_LANGUAGE = (
    "Candidate curve-mapping context question: provide matching source/plot "
    "columns, sample identity, and version context before evaluating point coverage."
)
DO_NOT_OVERCLAIM = (
    "This comparison reports a low-priority coverage question only; it does not "
    "infer why points differ or determine research misconduct."
)
ALTERNATIVE_EXPLANATIONS = [
    "The displayed axis limits may clip source points outside the plotted range.",
    "A disclosed downsampling rule may intentionally retain only a subset of points.",
    "Smoothing may change plotted y-values without changing source-data availability.",
    "NaN or non-finite source values may be omitted by plotting software.",
    "A disclosed filtering rule may exclude the identified interval.",
]
MANUAL_REQUESTS = [
    "Confirm the exact source-data and plot-data x/y column mapping.",
    "Provide plotting code or settings for axis limits, downsampling, smoothing, and filtering.",
    "Verify sample identity, source version, table locations, and file hashes.",
]
LIMITATIONS = [
    "Only supplied CSV/XLSX tables are compared.",
    "No plot image, screenshot, PDF figure, or raster curve is digitized.",
    "The v1 check evaluates x-axis point coverage and does not adjudicate plotting intent.",
]

SEGMENT_SAFE_LANGUAGE = (
    "Candidate segment-shape similarity between two explicitly mapped numeric "
    "curve segments; manual review of source, processing, sampling, and "
    "instrument context is required."
)
SEGMENT_DO_NOT_OVERCLAIM = (
    "This deterministic comparison reports a candidate segment-shape similarity "
    "only; it does not determine why the shapes align."
)
SEGMENT_ALTERNATIVE_EXPLANATIONS = [
    "Smoothing, interpolation, or resampling can align otherwise independent curve segments.",
    "Periodic measurements or repeated acquisition cycles can produce similar local shapes.",
    "A shared reference, baseline subtraction, or normalization can preserve the same shape.",
    "A documented offset or scale conversion can create an affine relationship between curves.",
    "Instrument quantization or limited export precision can increase apparent similarity.",
    "The two plotted series may intentionally derive from the same disclosed upstream data.",
]
SEGMENT_MANUAL_REQUESTS = [
    "Confirm that the mapped columns represent independently labelled curves and verify both source hashes.",
    "Inspect the original exports and plotting code for smoothing, interpolation, resampling, normalization, offsets, and scale conversions.",
    "Check sampling cadence, repeated or periodic acquisition cycles, and instrument/export precision.",
    "Verify the exact row and x-axis spans against the supplied source tables.",
]
SEGMENT_LIMITATIONS = [
    "Only explicitly mapped numeric columns in supplied CSV/XLSX tables are compared.",
    "No PDF, figure image, screenshot, raster curve, OCR, or model extraction is used.",
    "Constant, near-linear, low-dynamic-range, axis-inverted, overlapping, blank-row-bridged, and short windows are excluded.",
    "The bounded candidate search can be slow near its declared point budget; larger inputs require a reviewed subset or an explicit budget change.",
    "Affine shape similarity cannot identify its cause without manual source and processing review.",
]


class CurveReconciliationError(ValueError):
    """Raised when supplied structured curve tables cannot be reconciled safely."""


@dataclass(frozen=True)
class _SegmentMatch:
    first_start: int
    second_start: int
    window_length: int
    correlation: float
    normalized_rmse: float
    scale: float
    offset: float
    sampling_correlation: float
    sampling_normalized_rmse: float
    sampling_scale: float
    sampling_offset: float


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _as_decimal(value: Any) -> Decimal | None:
    if isinstance(value, bool):
        return None
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    return number if number.is_finite() else None


def _decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _file_hash(table: CurveTableSpec) -> str:
    digest = hashlib.sha256()
    try:
        with table.path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError as exc:
        raise CurveReconciliationError(
            f"unable to read supplied table {table.path.name}: {exc}"
        ) from exc
    computed = "sha256:" + digest.hexdigest()
    if table.source_hash and table.source_hash.lower() != computed:
        raise CurveReconciliationError(
            f"supplied source hash does not match table contents: {table.path.name}"
        )
    return computed


def _column_index(columns: Sequence[str], requested: str, *, table_name: str) -> int:
    if requested in columns:
        return columns.index(requested)
    target = requested.casefold()
    matches = [index for index, column in enumerate(columns) if column.casefold() == target]
    if len(matches) == 1:
        return matches[0]
    raise CurveReconciliationError(
        f"{table_name}: mapped column {requested!r} was not found unambiguously"
    )


def _positioned_rows(
    raw_rows: Sequence[Sequence[Any]],
) -> tuple[list[tuple[int, list[str]]], list[str], list[str]]:
    header_index = next(
        (
            index
            for index, row in enumerate(raw_rows)
            if any(str(cell).strip() for cell in row if cell is not None)
        ),
        None,
    )
    if header_index is None:
        return [], [], ["no table columns parsed"]
    columns = [
        str(cell).strip() if cell is not None else ""
        for cell in raw_rows[header_index]
    ]
    rows: list[tuple[int, list[str]]] = []
    for zero_based_index, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 1):
        rows.append(
            (
                zero_based_index + 1,
                [
                    str(cell).strip() if cell is not None else ""
                    for cell in raw_row
                ],
            )
        )
    return rows, columns, []


def _read_positioned_curve_rows(
    table: CurveTableSpec,
) -> tuple[list[tuple[int, list[str]]], list[str], list[str]]:
    if table.path.suffix.lower() == ".csv":
        try:
            content = table.path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            try:
                content = table.path.read_text(encoding="latin-1")
            except OSError as exc:
                return [], [], [f"unable to read table: {exc}"]
        except OSError as exc:
            return [], [], [f"unable to read table: {exc}"]
        try:
            raw_rows = list(csv.reader(io.StringIO(content)))
        except csv.Error as exc:
            return [], [], [f"CSV parsing error: {exc}"]
        return _positioned_rows(raw_rows)

    if table.path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            return [], [], ["openpyxl_not_installed"]
        workbook = None
        try:
            workbook = openpyxl.load_workbook(
                filename=str(table.path),
                data_only=True,
                read_only=True,
                keep_links=False,
            )
            selected_sheet = table.sheet_name or (
                workbook.sheetnames[0] if workbook.sheetnames else None
            )
            if selected_sheet is None:
                return [], [], ["Workbook has no worksheets."]
            if selected_sheet not in workbook.sheetnames:
                return [], [], [f"Sheet '{selected_sheet}' not found."]
            raw_rows = [
                list(row)
                for row in workbook[selected_sheet].iter_rows(values_only=True)
            ]
            return _positioned_rows(raw_rows)
        except Exception as exc:
            return [], [], [f"Failed to read sheet: {exc}"]
        finally:
            if workbook is not None:
                workbook.close()

    return [], [], [f"Unsupported curve table format: {table.path.suffix.lower()}"]


def _read_points(
    table: CurveTableSpec,
    *,
    x_column: str,
    y_column: str,
) -> list[CurvePoint]:
    if not table.path.is_file():
        raise CurveReconciliationError(f"supplied table not found: {table.path.name}")

    rows, columns, warnings = _read_positioned_curve_rows(table)
    if not columns:
        detail = "; ".join(warnings) or "no table columns parsed"
        raise CurveReconciliationError(f"{table.path.name}: {detail}")
    x_index = _column_index(columns, x_column, table_name=table.path.name)
    y_index = _column_index(columns, y_column, table_name=table.path.name)

    points: list[CurvePoint] = []
    for row_number, row in rows:
        raw_x = row[x_index] if x_index < len(row) else ""
        raw_y = row[y_index] if y_index < len(row) else ""
        x_value = _as_decimal(raw_x)
        if x_value is None:
            continue
        points.append(
            CurvePoint(
                x=x_value,
                y=_as_decimal(raw_y),
                row_number=row_number,
                sequence_index=row_number,
            )
        )
    if not points:
        raise CurveReconciliationError(
            f"{table.path.name}: mapped x column contains no finite numeric points"
        )
    return points


def _coerce_spec(spec: CurveReconciliationSpec | Mapping[str, Any]) -> CurveReconciliationSpec:
    if isinstance(spec, CurveReconciliationSpec):
        return spec
    if isinstance(spec, Mapping):
        return CurveReconciliationSpec.model_validate(dict(spec))
    raise TypeError("curve reconciliation requires a CurveReconciliationSpec or mapping")


def _x_matches(value: Decimal, candidates: Iterable[Decimal], tolerance: Decimal) -> bool:
    return any(abs(value - candidate) <= tolerance for candidate in candidates)


def _inside_interval(value: Decimal, start: Decimal, end: Decimal) -> bool:
    return start <= value <= end


def _group_contiguous_missing(
    missing: Sequence[CurvePoint],
    *,
    minimum_count: int,
) -> list[dict[str, Any]]:
    if not missing:
        return []
    groups: list[list[CurvePoint]] = [[missing[0]]]
    for point in missing[1:]:
        if point.sequence_index == groups[-1][-1].sequence_index + 1:
            groups[-1].append(point)
        else:
            groups.append([point])
    return [
        {
            "x_start": _decimal_text(group[0].x),
            "x_end": _decimal_text(group[-1].x),
            "point_count": len(group),
        }
        for group in groups
        if len(group) >= minimum_count
    ]


def _mapping_record(mapping: CurveColumnMapping | None) -> dict[str, Any] | None:
    return mapping.model_dump(mode="json") if mapping is not None else None


def _evidence_item(
    table: CurveTableSpec,
    *,
    table_role: str,
    mapping: CurveColumnMapping | None,
) -> EvidenceItem:
    if mapping is None:
        x_column = None
        y_column = None
    elif table_role == "source_data":
        x_column = mapping.source_x
        y_column = mapping.source_y
    else:
        x_column = mapping.plot_x
        y_column = mapping.plot_y
    return EvidenceItem(
        source=table.public_source,
        location=table.location,
        metadata={
            "table_role": table_role,
            "source_hash": _file_hash(table),
            "sample_id": table.sample_id,
            "source_version": table.source_version,
            "sheet_name": table.sheet_name,
            "x_column": x_column,
            "y_column": y_column,
        },
    )


def _finding_id(
    spec: CurveReconciliationSpec,
    *,
    comparison_kind: str,
    missing_intervals: Sequence[Mapping[str, Any]],
) -> str:
    payload = json.dumps(
        {
            "source": spec.source_table.public_source,
            "plot": spec.plot_table.public_source,
            "sample": spec.source_table.sample_id,
            "version": spec.source_table.source_version,
            "kind": comparison_kind,
            "intervals": list(missing_intervals),
        },
        sort_keys=True,
        ensure_ascii=False,
    )
    return "CURVE-" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:12].upper()


def _build_question(
    spec: CurveReconciliationSpec,
    *,
    comparison_kind: str,
    missing_intervals: Sequence[Mapping[str, Any]] = (),
    rule: DetectorRule | None = None,
) -> Finding:
    is_missing_interval = comparison_kind == "unexplained_contiguous_missing_interval"
    safe_language = (
        rule.safe_report_language
        if is_missing_interval and rule is not None
        else SAFE_MISSING_INTERVAL_LANGUAGE if is_missing_interval else SAFE_CONTEXT_LANGUAGE
    )
    mapping = spec.mapping
    return Finding(
        finding_id=_finding_id(
            spec,
            comparison_kind=comparison_kind,
            missing_intervals=missing_intervals,
        ),
        type="curve_point_coverage_question",
        title=(
            "Supplied-table curve point coverage question"
            if is_missing_interval
            else "Curve source/plot mapping context required"
        ),
        risk=RiskLevel.LOW,
        summary=(
            "A contiguous internal x-axis interval in the supplied source-data table is not represented in the supplied plot-data table."
            if is_missing_interval
            else "Point coverage cannot be evaluated until matching source and plot context is supplied."
        ),
        evidence=[
            _evidence_item(spec.source_table, table_role="source_data", mapping=mapping),
            _evidence_item(spec.plot_table, table_role="plot_data", mapping=mapping),
        ],
        manual_verification=ManualVerification(
            needed=True,
            requests=(
                list(rule.manual_verification)
                if rule is not None
                else list(MANUAL_REQUESTS)
            ),
        ),
        safe_report_language=safe_language,
        finding_category=RULE_ID,
        false_positive_risks=(
            list(rule.false_positive_risks)
            if rule is not None
            else list(ALTERNATIVE_EXPLANATIONS)
        ),
        alternative_explanations=list(ALTERNATIVE_EXPLANATIONS),
        limitations=list(LIMITATIONS),
        provenance={
            "rule_id": RULE_ID,
            "comparison_kind": comparison_kind,
            "comparison_mode": "structured_table_point_coverage",
            "automatic_digitization_performed": False,
            "mapping": _mapping_record(mapping),
            "missing_intervals": [dict(interval) for interval in missing_intervals],
            "sample_id": spec.source_table.sample_id,
            "source_version": spec.source_table.source_version,
            "source_versions": sorted(
                {
                    spec.source_table.source_version,
                    spec.plot_table.source_version,
                }
            ),
            "disclosure": spec.disclosure.model_dump(mode="json"),
            "open_for_scoring": False,
            "mrpi_eligible": False,
            "do_not_overclaim": DO_NOT_OVERCLAIM,
        },
    )


def reconcile_curve_coverage(
    spec: CurveReconciliationSpec | Mapping[str, Any],
    *,
    rule: DetectorRule | None = None,
) -> list[Finding]:
    """Compare x-axis coverage between supplied source and plot-data tables.

    This function never reads a figure image. Every output is a low,
    non-scoring verification question grounded in the two supplied tables.
    """
    normalized = _coerce_spec(spec)
    if normalized.mapping is None:
        return [
            _build_question(
                normalized,
                comparison_kind="missing_source_mapping",
                rule=rule,
            )
        ]
    if not normalized.context_matches:
        return [
            _build_question(
                normalized,
                comparison_kind="sample_or_version_context_mismatch",
                rule=rule,
            )
        ]
    if (
        normalized.disclosure.filtering_disclosed
        and not normalized.disclosure.filtered_intervals
    ):
        return [
            _build_question(
                normalized,
                comparison_kind="filtering_disclosed_without_intervals",
                rule=rule,
            )
        ]

    mapping = normalized.mapping
    source_points = _read_points(
        normalized.source_table,
        x_column=mapping.source_x,
        y_column=mapping.source_y,
    )
    plot_points = _read_points(
        normalized.plot_table,
        x_column=mapping.plot_x,
        y_column=mapping.plot_y,
    )
    usable_source = [point for point in source_points if point.y is not None]
    usable_plot = [point for point in plot_points if point.y is not None]
    if not usable_source or not usable_plot:
        return [
            _build_question(
                normalized,
                comparison_kind="insufficient_finite_curve_points",
                rule=rule,
            )
        ]

    plot_x = [point.x for point in usable_plot]
    source_x = [point.x for point in usable_source]
    if normalized.disclosure.downsampling_disclosed and all(
        _x_matches(value, source_x, mapping.x_tolerance) for value in plot_x
    ):
        return []

    lower_plot = min(plot_x)
    upper_plot = max(plot_x)
    axis_limits = normalized.disclosure.axis_limits
    filtered_intervals = normalized.disclosure.filtered_intervals
    missing: list[CurvePoint] = []
    for point in usable_source:
        if axis_limits is not None and not _inside_interval(
            point.x, axis_limits[0], axis_limits[1]
        ):
            continue
        if not (lower_plot < point.x < upper_plot):
            continue
        if _x_matches(point.x, plot_x, mapping.x_tolerance):
            continue
        if any(
            _inside_interval(point.x, interval.start, interval.end)
            for interval in filtered_intervals
        ):
            continue
        missing.append(point)

    intervals = _group_contiguous_missing(
        missing,
        minimum_count=mapping.minimum_contiguous_missing,
    )
    if not intervals:
        return []
    return [
        _build_question(
            normalized,
            comparison_kind="unexplained_contiguous_missing_interval",
            missing_intervals=intervals,
            rule=rule,
        )
    ]


def _dynamic_range(values: Sequence[float]) -> tuple[float, float]:
    spread = max(values) - min(values)
    magnitude = max(max(abs(value) for value in values), 1e-15)
    return spread, spread / magnitude


def _linear_r_squared(
    x_values: Sequence[float],
    y_values: Sequence[float],
) -> float:
    metrics = _affine_metrics(x_values, y_values)
    if metrics is None:
        return 1.0
    return metrics["correlation"] ** 2


def _strictly_monotonic(values: Sequence[float]) -> bool:
    if len(values) < 2:
        return False
    differences = [right - left for left, right in zip(values, values[1:])]
    return all(value > 0 for value in differences) or all(
        value < 0 for value in differences
    )


def _affine_metrics(
    first: Sequence[float],
    second: Sequence[float],
) -> dict[str, float] | None:
    if len(first) != len(second) or len(first) < 2:
        return None
    count = len(first)
    first_mean = sum(first) / count
    second_mean = sum(second) / count
    first_centered = [value - first_mean for value in first]
    second_centered = [value - second_mean for value in second]
    first_ss = sum(value * value for value in first_centered)
    second_ss = sum(value * value for value in second_centered)
    if first_ss <= 0 or second_ss <= 0:
        return None
    covariance = sum(
        left * right for left, right in zip(first_centered, second_centered)
    )
    scale = covariance / first_ss
    offset = second_mean - scale * first_mean
    residual_ss = sum(
        (right - (offset + scale * left)) ** 2
        for left, right in zip(first, second)
    )
    rmse = math.sqrt(residual_ss / count)
    second_range = max(second) - min(second)
    if second_range <= 0:
        return None
    correlation = covariance / math.sqrt(first_ss * second_ss)
    return {
        "correlation": max(-1.0, min(1.0, correlation)),
        "normalized_rmse": rmse / second_range,
        "scale": scale,
        "offset": offset,
    }


def _qualifying_segment_metrics(
    first_x: Sequence[float],
    first_y: Sequence[float],
    second_x: Sequence[float],
    second_y: Sequence[float],
    options: CurveSegmentSimilarityOptions,
) -> dict[str, float] | None:
    first_range, first_relative = _dynamic_range(first_y)
    second_range, second_relative = _dynamic_range(second_y)
    if (
        first_range < options.minimum_dynamic_range
        or second_range < options.minimum_dynamic_range
        or first_relative < options.minimum_relative_dynamic_range
        or second_relative < options.minimum_relative_dynamic_range
    ):
        return None
    if not _strictly_monotonic(first_x) or not _strictly_monotonic(second_x):
        return None
    sampling_metrics = _affine_metrics(first_x, second_x)
    if sampling_metrics is None:
        return None
    if (
        sampling_metrics["scale"] <= 0
        or sampling_metrics["correlation"]
        < options.sampling_correlation_threshold
        or sampling_metrics["normalized_rmse"]
        > options.sampling_normalized_rmse_threshold
    ):
        return None
    if (
        _linear_r_squared(first_x, first_y) >= options.near_linear_r2_threshold
        or _linear_r_squared(second_x, second_y) >= options.near_linear_r2_threshold
    ):
        return None
    metrics = _affine_metrics(first_y, second_y)
    if metrics is None:
        return None
    if (
        metrics["scale"] <= 0
        or metrics["correlation"] < options.correlation_threshold
        or metrics["normalized_rmse"] > options.normalized_rmse_threshold
    ):
        return None
    metrics["sampling_correlation"] = sampling_metrics["correlation"]
    metrics["sampling_normalized_rmse"] = sampling_metrics["normalized_rmse"]
    metrics["sampling_scale"] = sampling_metrics["scale"]
    metrics["sampling_offset"] = sampling_metrics["offset"]
    return metrics


def _resolved_sheet_identity(table: CurveTableSpec) -> str | None:
    if table.path.suffix.lower() != ".xlsx":
        return None
    if table.sheet_name is not None:
        return table.sheet_name.casefold()
    from integrity_agent.core.tables.adapters.xlsx_table import get_xlsx_sheets

    sheets = get_xlsx_sheets(table.path)
    return sheets[0].casefold() if sheets else None


def _normalized_column_identity(value: str) -> str:
    return " ".join(value.split()).casefold()


def _curve_series_identity(
    table: CurveTableSpec,
    *,
    x_column: str,
    y_column: str,
) -> str:
    return json.dumps(
        {
            "source": table.public_source,
            "source_hash": _file_hash(table),
            "sheet_name": _resolved_sheet_identity(table),
            "x_column": _normalized_column_identity(x_column),
            "y_column": _normalized_column_identity(y_column),
        },
        sort_keys=True,
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _same_curve_series(
    spec: CurveReconciliationSpec,
    mapping: CurveColumnMapping,
) -> bool:
    try:
        same_path = spec.source_table.path.resolve() == spec.plot_table.path.resolve()
    except OSError:
        same_path = spec.source_table.path == spec.plot_table.path
    if not same_path:
        return False
    suffix = spec.source_table.path.suffix.lower()
    if suffix == ".xlsx":
        if _resolved_sheet_identity(
            spec.source_table
        ) != _resolved_sheet_identity(spec.plot_table):
            return False

    return bool(
        _normalized_column_identity(mapping.source_x)
        == _normalized_column_identity(mapping.plot_x)
        and _normalized_column_identity(mapping.source_y)
        == _normalized_column_identity(mapping.plot_y)
    )


def _contiguous_run_lengths(points: Sequence[CurvePoint]) -> list[int]:
    lengths = [1] * len(points)
    for index in range(len(points) - 2, -1, -1):
        if points[index + 1].sequence_index == points[index].sequence_index + 1:
            lengths[index] = lengths[index + 1] + 1
    return lengths


def _find_best_segment_match(
    first_points: Sequence[CurvePoint],
    second_points: Sequence[CurvePoint],
    *,
    options: CurveSegmentSimilarityOptions,
    same_series: bool,
    first_series_identity: str,
    second_series_identity: str,
) -> _SegmentMatch | None:
    minimum = options.minimum_window_points
    if len(first_points) < minimum or len(second_points) < minimum:
        return None
    first_values = [float(point.y) for point in first_points if point.y is not None]
    second_values = [float(point.y) for point in second_points if point.y is not None]
    first_x = [float(point.x) for point in first_points]
    second_x = [float(point.x) for point in second_points]
    first_contiguous = _contiguous_run_lengths(first_points)
    second_contiguous = _contiguous_run_lengths(second_points)
    seeds: list[_SegmentMatch] = []

    for first_start in range(len(first_values) - minimum + 1):
        for second_start in range(len(second_values) - minimum + 1):
            if same_series and second_start <= first_start:
                continue
            maximum = min(
                len(first_values) - first_start,
                len(second_values) - second_start,
                first_contiguous[first_start],
                second_contiguous[second_start],
            )
            if same_series:
                maximum = min(maximum, second_start - first_start)
            if maximum < minimum:
                continue

            metrics = _qualifying_segment_metrics(
                first_x[first_start : first_start + minimum],
                first_values[first_start : first_start + minimum],
                second_x[second_start : second_start + minimum],
                second_values[second_start : second_start + minimum],
                options,
            )
            if metrics is None:
                continue

            seeds.append(
                _SegmentMatch(
                    first_start=first_start,
                    second_start=second_start,
                    window_length=minimum,
                    correlation=metrics["correlation"],
                    normalized_rmse=metrics["normalized_rmse"],
                    scale=metrics["scale"],
                    offset=metrics["offset"],
                    sampling_correlation=metrics["sampling_correlation"],
                    sampling_normalized_rmse=metrics[
                        "sampling_normalized_rmse"
                    ],
                    sampling_scale=metrics["sampling_scale"],
                    sampling_offset=metrics["sampling_offset"],
                )
            )

    if not seeds:
        return None
    def canonical_match_identity(item: _SegmentMatch) -> str:
        descriptors = []
        for identity, points, start in (
            (first_series_identity, first_points, item.first_start),
            (second_series_identity, second_points, item.second_start),
        ):
            segment = points[start : start + item.window_length]
            descriptors.append(
                {
                    "series": identity,
                    "row_start": segment[0].row_number,
                    "row_end": segment[-1].row_number,
                    "x_start": _decimal_text(segment[0].x),
                    "x_end": _decimal_text(segment[-1].x),
                }
            )
        descriptors.sort(
            key=lambda descriptor: json.dumps(
                descriptor,
                sort_keys=True,
                ensure_ascii=False,
            )
        )
        return json.dumps(
            descriptors,
            sort_keys=True,
            ensure_ascii=False,
            separators=(",", ":"),
        )

    selected_seeds = sorted(
        seeds,
        key=lambda item: (
            -round(item.correlation, 12),
            canonical_match_identity(item),
            round(item.normalized_rmse, 12),
        ),
    )[: options.maximum_seed_candidates]

    matches: list[_SegmentMatch] = []
    for seed in selected_seeds:
        best = seed
        maximum = min(
            len(first_values) - seed.first_start,
            len(second_values) - seed.second_start,
            first_contiguous[seed.first_start],
            second_contiguous[seed.second_start],
        )
        if same_series:
            maximum = min(maximum, seed.second_start - seed.first_start)

        for candidate_length in range(minimum + 1, maximum + 1):
            metrics = _qualifying_segment_metrics(
                first_x[seed.first_start : seed.first_start + candidate_length],
                first_values[
                    seed.first_start : seed.first_start + candidate_length
                ],
                second_x[
                    seed.second_start : seed.second_start + candidate_length
                ],
                second_values[
                    seed.second_start : seed.second_start + candidate_length
                ],
                options,
            )
            if metrics is not None:
                best = _SegmentMatch(
                    first_start=seed.first_start,
                    second_start=seed.second_start,
                    window_length=candidate_length,
                    correlation=metrics["correlation"],
                    normalized_rmse=metrics["normalized_rmse"],
                    scale=metrics["scale"],
                    offset=metrics["offset"],
                    sampling_correlation=metrics["sampling_correlation"],
                    sampling_normalized_rmse=metrics[
                        "sampling_normalized_rmse"
                    ],
                    sampling_scale=metrics["sampling_scale"],
                    sampling_offset=metrics["sampling_offset"],
                )

        expanded_right = best
        first_end = expanded_right.first_start + expanded_right.window_length
        second_end = expanded_right.second_start + expanded_right.window_length
        maximum_left = min(
            expanded_right.first_start,
            expanded_right.second_start,
        )
        for left_extension in range(1, maximum_left + 1):
            first_start = expanded_right.first_start - left_extension
            second_start = expanded_right.second_start - left_extension
            candidate_length = expanded_right.window_length + left_extension
            if (
                first_contiguous[first_start] < candidate_length
                or second_contiguous[second_start] < candidate_length
            ):
                continue
            if same_series and not (
                first_end <= second_start or second_end <= first_start
            ):
                continue
            metrics = _qualifying_segment_metrics(
                first_x[first_start:first_end],
                first_values[first_start:first_end],
                second_x[second_start:second_end],
                second_values[second_start:second_end],
                options,
            )
            if metrics is not None:
                best = _SegmentMatch(
                    first_start=first_start,
                    second_start=second_start,
                    window_length=candidate_length,
                    correlation=metrics["correlation"],
                    normalized_rmse=metrics["normalized_rmse"],
                    scale=metrics["scale"],
                    offset=metrics["offset"],
                    sampling_correlation=metrics["sampling_correlation"],
                    sampling_normalized_rmse=metrics[
                        "sampling_normalized_rmse"
                    ],
                    sampling_scale=metrics["sampling_scale"],
                    sampling_offset=metrics["sampling_offset"],
                )
        matches.append(best)

    return min(
        matches,
        key=lambda item: (
            -item.window_length,
            -round(item.correlation, 12),
            canonical_match_identity(item),
            round(item.normalized_rmse, 12),
        ),
    )


def _segment_record(
    points: Sequence[CurvePoint],
    *,
    start: int,
    length: int,
) -> dict[str, Any]:
    segment = points[start : start + length]
    return {
        "row_start": segment[0].row_number,
        "row_end": segment[-1].row_number,
        "x_start": _decimal_text(segment[0].x),
        "x_end": _decimal_text(segment[-1].x),
    }


def _segment_evidence_item(
    table: CurveTableSpec,
    *,
    role: str,
    mapping: CurveColumnMapping,
    segment: Mapping[str, Any],
) -> EvidenceItem:
    if role == "first_segment":
        x_column = mapping.source_x
        y_column = mapping.source_y
    else:
        x_column = mapping.plot_x
        y_column = mapping.plot_y
    return EvidenceItem(
        source=table.public_source,
        location=(
            f"{table.location}; rows {segment['row_start']}-{segment['row_end']}; "
            f"x {segment['x_start']} to {segment['x_end']}"
        ),
        metadata={
            "segment_role": role,
            "source_hash": _file_hash(table),
            "sample_id": table.sample_id,
            "source_version": table.source_version,
            "sheet_name": table.sheet_name,
            "x_column": x_column,
            "y_column": y_column,
            **dict(segment),
        },
    )


def _segment_finding_id(
    spec: CurveReconciliationSpec,
    mapping: CurveColumnMapping,
    first_segment: Mapping[str, Any],
    second_segment: Mapping[str, Any],
) -> str:
    segments = [
        {
            "source": spec.source_table.public_source,
            "source_hash": _file_hash(spec.source_table),
            "sheet_name": _resolved_sheet_identity(spec.source_table),
            "x_column": _normalized_column_identity(mapping.source_x),
            "y_column": _normalized_column_identity(mapping.source_y),
            "segment": dict(first_segment),
        },
        {
            "source": spec.plot_table.public_source,
            "source_hash": _file_hash(spec.plot_table),
            "sheet_name": _resolved_sheet_identity(spec.plot_table),
            "x_column": _normalized_column_identity(mapping.plot_x),
            "y_column": _normalized_column_identity(mapping.plot_y),
            "segment": dict(second_segment),
        },
    ]
    segments.sort(
        key=lambda item: json.dumps(item, sort_keys=True, ensure_ascii=False)
    )
    payload = json.dumps(
        {"segments": segments},
        sort_keys=True,
        ensure_ascii=False,
    )
    return "CURVE-SEG-" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:12].upper()


def reconcile_curve_segment_similarity(
    spec: CurveReconciliationSpec | Mapping[str, Any],
    *,
    rule: DetectorRule | None = None,
) -> list[Finding]:
    """Compare explicitly mapped numeric curves for one best nontrivial segment match."""

    normalized = _coerce_spec(spec)
    options = normalized.segment_similarity
    mapping = normalized.mapping
    if options is None:
        return []
    if mapping is None:  # The schema rejects this; keep a defensive runtime guard.
        raise CurveReconciliationError(
            "segment similarity requires an explicit curve column mapping"
        )
    if _same_curve_series(normalized, mapping):
        return []

    first_series_identity = _curve_series_identity(
        normalized.source_table,
        x_column=mapping.source_x,
        y_column=mapping.source_y,
    )
    second_series_identity = _curve_series_identity(
        normalized.plot_table,
        x_column=mapping.plot_x,
        y_column=mapping.plot_y,
    )

    first_points = [
        point
        for point in _read_points(
            normalized.source_table,
            x_column=mapping.source_x,
            y_column=mapping.source_y,
        )
        if point.y is not None
    ]
    second_points = [
        point
        for point in _read_points(
            normalized.plot_table,
            x_column=mapping.plot_x,
            y_column=mapping.plot_y,
        )
        if point.y is not None
    ]
    if (
        len(first_points) > options.maximum_points_per_curve
        or len(second_points) > options.maximum_points_per_curve
    ):
        raise CurveReconciliationError(
            "segment similarity input exceeds maximum_points_per_curve; "
            "supply a reviewed, package-relative subset or raise the explicit budget"
        )
    match = _find_best_segment_match(
        first_points,
        second_points,
        options=options,
        same_series=False,
        first_series_identity=first_series_identity,
        second_series_identity=second_series_identity,
    )
    if match is None:
        return []

    first_segment = _segment_record(
        first_points,
        start=match.first_start,
        length=match.window_length,
    )
    second_segment = _segment_record(
        second_points,
        start=match.second_start,
        length=match.window_length,
    )
    safe_language = rule.safe_report_language if rule is not None else SEGMENT_SAFE_LANGUAGE
    manual_requests = (
        list(rule.manual_verification)
        if rule is not None
        else list(SEGMENT_MANUAL_REQUESTS)
    )
    false_positive_risks = (
        list(rule.false_positive_risks)
        if rule is not None
        else list(SEGMENT_ALTERNATIVE_EXPLANATIONS)
    )
    correlation_sources = sorted(
        {
            normalized.source_table.public_source,
            normalized.plot_table.public_source,
        }
    )
    correlation_payload = json.dumps(
        correlation_sources,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    correlation_group = (
        "curve_segment_shape_similarity:"
        + hashlib.sha256(correlation_payload.encode("utf-8")).hexdigest()[:16]
    )
    return [
        Finding(
            finding_id=_segment_finding_id(
                normalized,
                mapping,
                first_segment,
                second_segment,
            ),
            type="curve_segment_shape_similarity_candidate",
            title="Candidate curve segment-shape similarity",
            risk=RiskLevel.MEDIUM,
            summary=(
                "Two explicitly mapped numeric curve segments meet the configured "
                "affine shape-similarity thresholds after excluding trivial windows."
            ),
            evidence=[
                _segment_evidence_item(
                    normalized.source_table,
                    role="first_segment",
                    mapping=mapping,
                    segment=first_segment,
                ),
                _segment_evidence_item(
                    normalized.plot_table,
                    role="second_segment",
                    mapping=mapping,
                    segment=second_segment,
                ),
            ],
            manual_verification=ManualVerification(
                needed=True,
                requests=manual_requests,
            ),
            safe_report_language=safe_language,
            finding_category=SEGMENT_SIMILARITY_RULE_ID,
            false_positive_risks=false_positive_risks,
            alternative_explanations=list(SEGMENT_ALTERNATIVE_EXPLANATIONS),
            limitations=list(SEGMENT_LIMITATIONS),
            provenance={
                "rule_id": SEGMENT_SIMILARITY_RULE_ID,
                "comparison_kind": "candidate_segment_shape_similarity",
                "comparison_mode": "supplied_numeric_curve_segments",
                "automatic_digitization_performed": False,
                "human_confirmed_independent_curves": True,
                "needs_manual_review": True,
                "window_length": match.window_length,
                "first_segment": first_segment,
                "second_segment": second_segment,
                "transform": {
                    "kind": "affine_offset_scale",
                    "scale": match.scale,
                    "offset": match.offset,
                },
                "similarity_metrics": {
                    "correlation": match.correlation,
                    "absolute_correlation": abs(match.correlation),
                    "normalized_rmse": match.normalized_rmse,
                    "correlation_threshold": options.correlation_threshold,
                    "normalized_rmse_threshold": options.normalized_rmse_threshold,
                },
                "sampling_metrics": {
                    "correlation": match.sampling_correlation,
                    "normalized_rmse": match.sampling_normalized_rmse,
                    "scale": match.sampling_scale,
                    "offset": match.sampling_offset,
                    "correlation_threshold": (
                        options.sampling_correlation_threshold
                    ),
                    "normalized_rmse_threshold": (
                        options.sampling_normalized_rmse_threshold
                    ),
                },
                "search_budget": {
                    "maximum_points_per_curve": options.maximum_points_per_curve,
                    "maximum_seed_candidates": options.maximum_seed_candidates,
                },
                "sample_ids": [
                    normalized.source_table.sample_id,
                    normalized.plot_table.sample_id,
                ],
                "source_versions": [
                    normalized.source_table.source_version,
                    normalized.plot_table.source_version,
                ],
                "open_for_scoring": True,
                "mrpi_eligible": True,
                "resolution_status": "open",
                "method_family": "curve_shape_similarity",
                "correlation_sources": correlation_sources,
                "correlation_group": correlation_group,
                "do_not_overclaim": SEGMENT_DO_NOT_OVERCLAIM,
            },
        )
    ]


reconcile_curve_points = reconcile_curve_coverage


def _write_summary(path: Path, *, findings: Sequence[Finding]) -> None:
    lines = [
        "# Curve Source Reconciliation Summary",
        "",
        "- Comparison mode: supplied CSV/XLSX source table to supplied plot-data table",
        "- Image digitization performed: no",
        "- Network used: no",
        f"- Low review questions: {len(findings)}",
        "- Open scoring findings: 0",
        "",
        "## Results",
    ]
    if findings:
        for finding in findings:
            lines.append(
                f"- `{finding.finding_id}`: {finding.safe_report_language}"
            )
    else:
        lines.append("- No unexplained contiguous internal coverage interval was identified.")
    lines.extend(
        [
            "",
            "## Do-not-overclaim notice",
            "- Coverage questions require plotting code, disclosures, and source mapping for interpretation.",
            "- The workflow does not infer intent or determine research misconduct.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def run_curve_reconciliation(
    spec: CurveReconciliationSpec | Mapping[str, Any],
    output_dir: Path | str = DEFAULT_OUTPUT_DIR,
) -> tuple[Path, Path]:
    normalized = _coerce_spec(spec)
    rule = load_rule_registry(_project_root() / "knowledge_base" / "detector_rules")[
        RULE_ID
    ]
    findings = reconcile_curve_coverage(normalized, rule=rule)
    records = [finding.to_ledger_record() for finding in findings]

    resolved_output = Path(output_dir)
    resolved_output.mkdir(parents=True, exist_ok=True)
    findings_path = resolved_output / FINDINGS_NAME
    summary_path = resolved_output / SUMMARY_NAME
    findings_tmp = findings_path.with_suffix(findings_path.suffix + ".tmp")
    summary_tmp = summary_path.with_suffix(summary_path.suffix + ".tmp")
    with findings_tmp.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
    validation = validate_ledger_file(findings_tmp)
    if not validation.ok:
        findings_tmp.unlink(missing_ok=True)
        details = "; ".join(issue.format() for issue in validation.issues)
        raise CurveReconciliationError(f"curve ledger validation failed: {details}")
    _write_summary(summary_tmp, findings=findings)
    findings_tmp.replace(findings_path)
    summary_tmp.replace(summary_path)
    return findings_path.resolve(), summary_path.resolve()


def run_curve_point_coverage(
    package_dir: Path,
    rule: DetectorRule,
    options: dict[str, Any] | None = None,
) -> list[RuleExecutionResult]:
    options = options or {}
    supplied_spec = options.get("spec") or options.get("curve_spec")
    if supplied_spec is None:
        return []
    findings = reconcile_curve_coverage(supplied_spec, rule=rule)
    results: list[RuleExecutionResult] = []
    for finding in findings:
        ledger = finding.to_ledger_record()
        manual = ledger["manual_verification"]
        results.append(
            RuleExecutionResult(
                finding_id=finding.finding_id,
                rule_id=rule.rule_id,
                risk_level=finding.risk.value,
                evidence_items=list(ledger["evidence"]),
                manual_verification=dict(manual),
                false_positive_risks=[str(item) for item in finding.false_positive_risks],
                safe_report_language=str(finding.safe_report_language),
                alternative_explanations=[
                    str(item) for item in finding.alternative_explanations
                ],
                missing_verification_materials=list(manual["requests"]),
                suggested_verification_questions=list(manual["requests"]),
                limitations=[str(item) for item in finding.limitations],
                metadata=dict(finding.provenance),
            )
        )
    return results


def run_curve_segment_similarity(
    package_dir: Path,
    rule: DetectorRule,
    options: dict[str, Any] | None = None,
) -> list[RuleExecutionResult]:
    options = options or {}
    supplied_spec = options.get("spec") or options.get("curve_spec")
    if supplied_spec is None:
        return []
    findings = reconcile_curve_segment_similarity(supplied_spec, rule=rule)
    results: list[RuleExecutionResult] = []
    for finding in findings:
        ledger = finding.to_ledger_record()
        manual = ledger["manual_verification"]
        results.append(
            RuleExecutionResult(
                finding_id=finding.finding_id,
                rule_id=rule.rule_id,
                risk_level=finding.risk.value,
                evidence_items=list(ledger["evidence"]),
                manual_verification=dict(manual),
                false_positive_risks=[
                    str(item) for item in finding.false_positive_risks
                ],
                safe_report_language=str(finding.safe_report_language),
                alternative_explanations=[
                    str(item) for item in finding.alternative_explanations
                ],
                missing_verification_materials=list(manual["requests"]),
                suggested_verification_questions=list(manual["requests"]),
                limitations=[str(item) for item in finding.limitations],
                metadata=dict(finding.provenance),
            )
        )
    return results


__all__ = [
    "CurveReconciliationError",
    "reconcile_curve_coverage",
    "reconcile_curve_points",
    "reconcile_curve_segment_similarity",
    "run_curve_point_coverage",
    "run_curve_reconciliation",
    "run_curve_segment_similarity",
]
