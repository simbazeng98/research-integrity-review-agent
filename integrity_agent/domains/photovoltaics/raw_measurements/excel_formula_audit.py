from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any
import openpyxl
from integrity_agent.domains.photovoltaics.raw_measurements.schema import ExcelFormulaAuditItem

def evaluate_simple_formula(formula_str: str, sheet_values: dict[str, Any]) -> float | None:
    # Remove leading '='
    expr = formula_str[1:].strip()
    
    # Check if formula contains only cells, numbers, basic operators (+ - * /) and parentheses
    # Cell format: A1, B12, AA3, etc.
    # Exclude formulas with sheet references (e.g. Sheet2!A1), functions (e.g. SUM), or brackets []
    if any(c in expr for c in ("!", "[", "]", ":", "$")):
        return None
        
    # Check if contains any letters that are not cell references
    # Cell references are like AA12. Let's find all word tokens.
    tokens = re.findall(r"[A-Za-z0-9_.]+", expr)
    cell_pattern = re.compile(r"^[A-Z]+[0-9]+$", re.IGNORECASE)
    
    for tok in tokens:
        # If token is numeric or coordinate, it is okay
        try:
            float(tok)
            continue
        except ValueError:
            pass
            
        if cell_pattern.match(tok):
            continue
            
        # Any other word token means it's a function or something else
        return None

    # Replace cell references with their numeric values
    # Sort keys by length descending to avoid replacing A10 before A1
    coords = [tok for tok in tokens if cell_pattern.match(tok)]
    coords.sort(key=len, reverse=True)
    
    for coord in coords:
        val = sheet_values.get(coord.upper())
        if val is None:
            return None
        try:
            numeric_val = float(val)
            # Replace coordinate with numeric value
            # Match coordinate as whole word to avoid partial replacement (e.g. replacing A1 in A10)
            expr = re.sub(rf"\b{coord}\b", str(numeric_val), expr, flags=re.IGNORECASE)
        except (ValueError, TypeError):
            return None

    # Sanity check: expression must only contain numbers, basic operators, and spaces
    # Allow float formatting (e.g. 1.2e-3)
    allowed_chars = set("0123456789.+-*/() eE")
    if not all(c in allowed_chars for c in expr):
        return None

    # Safe eval of the simple math expression
    try:
        # Use python eval with limited globals/locals to prevent arbitrary code execution
        res = eval(expr, {"__builtins__": None}, {})
        return float(res)
    except Exception:
        return None

def audit_excel_formulas(path: str) -> list[ExcelFormulaAuditItem]:
    filepath = Path(path)
    filename = filepath.name
    audit_items = []
    
    # 1. xlsm check
    if filepath.suffix.lower() == ".xlsm":
        audit_items.append(ExcelFormulaAuditItem(
            audit_id="audit-xlsm-refuse",
            source_file=filename,
            sheet_name="all",
            cell_coordinate="none",
            formula=None,
            cached_value=None,
            cell_value=None,
            audit_type="xlsm_not_supported_for_formula_audit_safety",
            message="Macro-enabled workbook (.xlsm) is not supported for formula audit safety.",
            severity="medium"
        ))
        return audit_items

    if not filepath.exists() or filepath.is_dir():
        return audit_items

    # Load formula workbook (data_only=False)
    try:
        wb_formula = openpyxl.load_workbook(filepath, data_only=False, read_only=False, keep_vba=False)
    except Exception as e:
        audit_items.append(ExcelFormulaAuditItem(
            audit_id="audit-load-error",
            source_file=filename,
            sheet_name="all",
            cell_coordinate="none",
            formula=None,
            cached_value=None,
            cell_value=None,
            audit_type="load_error",
            message=f"Failed to load workbook formulas: {e}",
            severity="medium"
        ))
        return audit_items

    # Load cached values workbook (data_only=True)
    try:
        wb_data = openpyxl.load_workbook(filepath, data_only=True, read_only=False, keep_vba=False)
    except Exception as e:
        audit_items.append(ExcelFormulaAuditItem(
            audit_id="audit-load-error-data",
            source_file=filename,
            sheet_name="all",
            cell_coordinate="none",
            formula=None,
            cached_value=None,
            cell_value=None,
            audit_type="load_error",
            message=f"Failed to load workbook cached values: {e}",
            severity="medium"
        ))
        # fallback to continue with just formula
        wb_data = None

    audit_counter = 1

    for sheet_name in wb_formula.sheetnames:
        ws_form = wb_formula[sheet_name]
        ws_data = wb_data[sheet_name] if wb_data else None
        
        # Check sheet visibility
        if ws_form.sheet_state != "visible":
            audit_items.append(ExcelFormulaAuditItem(
                audit_id=f"audit-{audit_counter:04d}",
                source_file=filename,
                sheet_name=sheet_name,
                cell_coordinate="none",
                formula=None,
                cached_value=None,
                cell_value=None,
                audit_type="hidden_sheet",
                message=f"Sheet '{sheet_name}' is hidden (state: {ws_form.sheet_state}).",
                severity="low"
            ))
            audit_counter += 1

        # Check hidden columns/rows
        for row_idx, row_dim in ws_form.row_dimensions.items():
            if row_dim.hidden:
                audit_items.append(ExcelFormulaAuditItem(
                    audit_id=f"audit-{audit_counter:04d}",
                    source_file=filename,
                    sheet_name=sheet_name,
                    cell_coordinate=f"Row {row_idx}",
                    formula=None,
                    cached_value=None,
                    cell_value=None,
                    audit_type="hidden_row",
                    message=f"Row {row_idx} is hidden in sheet '{sheet_name}'.",
                    severity="low"
                ))
                audit_counter += 1
                
        for col_letter, col_dim in ws_form.column_dimensions.items():
            if col_dim.hidden:
                audit_items.append(ExcelFormulaAuditItem(
                    audit_id=f"audit-{audit_counter:04d}",
                    source_file=filename,
                    sheet_name=sheet_name,
                    cell_coordinate=f"Col {col_letter}",
                    formula=None,
                    cached_value=None,
                    cell_value=None,
                    audit_type="hidden_column",
                    message=f"Column {col_letter} is hidden in sheet '{sheet_name}'.",
                    severity="low"
                ))
                audit_counter += 1

        # Collect cell values for evaluation
        sheet_values = {}
        if ws_data:
            for r in range(1, ws_data.max_row + 1):
                for c in range(1, ws_data.max_column + 1):
                    val = ws_data.cell(row=r, column=c).value
                    coord = ws_data.cell(row=r, column=c).coordinate
                    sheet_values[coord.upper()] = val

        # Now iterate cells to find formulas
        for r in range(1, ws_form.max_row + 1):
            # Check if this row is mostly formulas
            # We will gather info for hardcoded output detection
            row_cells_form = []
            row_cells_numeric_hardcoded = []
            
            for c in range(1, ws_form.max_column + 1):
                cell_f = ws_form.cell(row=r, column=c)
                coord = cell_f.coordinate
                val_f = cell_f.value
                val_d = ws_data.cell(row=r, column=c).value if ws_data else None
                
                is_formula = isinstance(val_f, str) and val_f.startswith("=")
                
                # Check for volatile functions or external references
                if is_formula:
                    val_f_upper = val_f.upper()
                    # Volatile checks
                    volatile_funcs = ["NOW", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT"]
                    for vf in volatile_funcs:
                        if f"{vf}(" in val_f_upper:
                            audit_items.append(ExcelFormulaAuditItem(
                                audit_id=f"audit-{audit_counter:04d}",
                                source_file=filename,
                                sheet_name=sheet_name,
                                cell_coordinate=coord,
                                formula=val_f,
                                cached_value=val_d,
                                cell_value=val_f,
                                audit_type="volatile_function",
                                message=f"Volatile function {vf} detected in formula.",
                                severity="low"
                            ))
                            audit_counter += 1
                    
                    # External check
                    if "[" in val_f or "]" in val_f:
                        audit_items.append(ExcelFormulaAuditItem(
                            audit_id=f"audit-{audit_counter:04d}",
                            source_file=filename,
                            sheet_name=sheet_name,
                            cell_coordinate=coord,
                            formula=val_f,
                            cached_value=val_d,
                            cell_value=val_f,
                            audit_type="external_reference",
                            message="External workbook reference detected in formula.",
                            severity="low"
                        ))
                        audit_counter += 1
                        
                    # Save for formula cell logging
                    audit_items.append(ExcelFormulaAuditItem(
                        audit_id=f"audit-{audit_counter:04d}",
                        source_file=filename,
                        sheet_name=sheet_name,
                        cell_coordinate=coord,
                        formula=val_f,
                        cached_value=val_d,
                        cell_value=val_f,
                        audit_type="formula_cell",
                        message=f"Formula cell: {val_f}",
                        severity="low"
                    ))
                    audit_counter += 1
                    row_cells_form.append(coord)

                    # Simple evaluation check
                    recalc_val = evaluate_simple_formula(val_f, sheet_values)
                    if recalc_val is not None and val_d is not None:
                        try:
                            d_val = float(val_d)
                            # Compare absolute and relative difference
                            abs_diff = abs(recalc_val - d_val)
                            rel_diff = abs_diff / abs(d_val) if d_val != 0 else 0
                            if abs_diff > 0.05 and rel_diff > 0.05:
                                audit_items.append(ExcelFormulaAuditItem(
                                    audit_id=f"audit-{audit_counter:04d}",
                                    source_file=filename,
                                    sheet_name=sheet_name,
                                    cell_coordinate=coord,
                                    formula=val_f,
                                    cached_value=val_d,
                                    cell_value=val_f,
                                    audit_type="formula_value_mismatch",
                                    message=f"Formula value mismatch: recomputed value {recalc_val:.4f} differs from cached value {val_d} in cell {coord}.",
                                    severity="medium"
                                ))
                                audit_counter += 1
                        except (ValueError, TypeError):
                            pass
                else:
                    # Not a formula. If numeric, track for hardcoded output detection
                    if val_f is not None:
                        try:
                            float(val_f)
                            row_cells_numeric_hardcoded.append((coord, val_f))
                        except (ValueError, TypeError):
                            pass

            # Hardcoded output detection within rows
            # If a row has many formulas but one hardcoded numeric value in a metric region
            if len(row_cells_form) >= 2 and len(row_cells_numeric_hardcoded) == 1:
                coord_hc, val_hc = row_cells_numeric_hardcoded[0]
                # check column headers of the row to see if it is a metric region
                # (For simplicity: if neighboring row/column cell contains formula, flag it)
                audit_items.append(ExcelFormulaAuditItem(
                    audit_id=f"audit-{audit_counter:04d}",
                    source_file=filename,
                    sheet_name=sheet_name,
                    cell_coordinate=coord_hc,
                    formula=None,
                    cached_value=val_hc,
                    cell_value=val_hc,
                    audit_type="hardcoded_output",
                    message=f"Hardcoded cell {coord_hc} containing numeric value {val_hc} found in row with formula cells.",
                    severity="medium"
                ))
                audit_counter += 1

        # Also do column-based check
        for c in range(1, ws_form.max_column + 1):
            col_cells_form = []
            col_cells_numeric_hardcoded = []
            col_header = ""
            
            # Find column header (usually row 1)
            first_cell_val = ws_form.cell(row=1, column=c).value
            if isinstance(first_cell_val, str):
                col_header = first_cell_val.lower()

            for r in range(1, ws_form.max_row + 1):
                cell_f = ws_form.cell(row=r, column=c)
                coord = cell_f.coordinate
                val_f = cell_f.value
                is_formula = isinstance(val_f, str) and val_f.startswith("=")
                
                if is_formula:
                    col_cells_form.append(coord)
                else:
                    if val_f is not None:
                        try:
                            float(val_f)
                            col_cells_numeric_hardcoded.append((coord, val_f))
                        except (ValueError, TypeError):
                            pass

            # Check hardcoded column pattern
            if len(col_cells_form) >= 2 and len(col_cells_numeric_hardcoded) == 1:
                coord_hc, val_hc = col_cells_numeric_hardcoded[0]
                audit_items.append(ExcelFormulaAuditItem(
                    audit_id=f"audit-{audit_counter:04d}",
                    source_file=filename,
                    sheet_name=sheet_name,
                    cell_coordinate=coord_hc,
                    formula=None,
                    cached_value=val_hc,
                    cell_value=val_hc,
                    audit_type="hardcoded_output",
                    message=f"Hardcoded cell {coord_hc} containing numeric value {val_hc} found in column with formula cells.",
                    severity="medium"
                ))
                audit_counter += 1

            # Check formula overwrite pattern (header matches metric, cell is hardcoded)
            # Apply to cells in row > 1 (excluding header)
            if any(k in col_header for k in ("pce", "eff", "calculated", "recompute")):
                for coord_hc, val_hc in col_cells_numeric_hardcoded:
                    # check if the coordinate is not the header row itself
                    cell_row = int(re.sub(r"\D", "", coord_hc))
                    if cell_row > 1:
                        audit_items.append(ExcelFormulaAuditItem(
                            audit_id=f"audit-{audit_counter:04d}",
                            source_file=filename,
                            sheet_name=sheet_name,
                            cell_coordinate=coord_hc,
                            formula=None,
                            cached_value=val_hc,
                            cell_value=val_hc,
                            audit_type="formula_overwrite_pattern",
                            message=f"Hardcoded numeric cell {coord_hc} found under calculated column header '{first_cell_val}'.",
                            severity="medium"
                        ))
                        audit_counter += 1

    try:
        wb_formula.close()
        if wb_data:
            wb_data.close()
    except Exception:
        pass

    return audit_items
