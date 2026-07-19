from __future__ import annotations

from pathlib import Path


def get_xlsx_sheets(file_path: Path | str) -> list[str]:
    """Get the sheet names in an Excel workbook."""
    file_path = Path(file_path)
    if file_path.suffix.lower() in [".xlsm", ".xls"]:
        return []

    try:
        import openpyxl
        # Read only sheetnames quickly
        wb = openpyxl.load_workbook(filename=str(file_path), read_only=True, keep_links=False)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []


def parse_xlsx_sheet(
    file_path: Path | str,
    sheet_name: str | None,
) -> tuple[list[list[str]], list[str], list[str]]:
    """Parse a single Excel sheet and return (rows_data, column_headers, warnings)."""
    file_path = Path(file_path)
    warnings: list[str] = []

    if file_path.suffix.lower() in [".xlsm", ".xls"]:
        warnings.append("xlsm_not_supported_for_safety")
        return [], [], warnings

    try:
        import openpyxl
    except ImportError:
        warnings.append("openpyxl_not_installed")
        return [], [], warnings

    wb = None
    try:
        wb = openpyxl.load_workbook(filename=str(file_path), data_only=True, read_only=True, keep_links=False)
        selected_sheet = sheet_name or (wb.sheetnames[0] if wb.sheetnames else None)
        if selected_sheet is None:
            warnings.append("Workbook has no worksheets.")
            return [], [], warnings
        if selected_sheet not in wb.sheetnames:
            warnings.append(f"Sheet '{selected_sheet}' not found.")
            return [], [], warnings

        ws = wb[selected_sheet]
        all_rows = []
        for r in ws.iter_rows(values_only=True):
            # Check if entire row is empty. If so, skip it.
            if any(cell is not None for cell in r):
                all_rows.append(r)

        if not all_rows:
            warnings.append("Empty sheet.")
            return [], [], warnings

        # Determine header and strip spaces
        columns = [str(c).strip() if c is not None else "" for c in all_rows[0]]

        rows_data = []
        for r in all_rows[1:]:
            row_vals = []
            # Make sure we pad or truncate to match header columns length
            for idx in range(len(columns)):
                val = r[idx] if idx < len(r) else None
                row_vals.append(str(val).strip() if val is not None else "")
            rows_data.append(row_vals)

        return rows_data, columns, warnings

    except Exception as e:
        warnings.append(f"Failed to read sheet: {e}")
        return [], [], warnings
    finally:
        if wb is not None:
            wb.close()
